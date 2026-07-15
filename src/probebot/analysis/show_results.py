"""
Probebot Show-Results — Trading Bot Evaluation Modes 1-4

Mode 1: OOS-Backtest (30% Test-Periode) aller Configs → Tabelle
Mode 2: Portfolio-Simulation (mehrere Configs zusammen)
Mode 3: Auto-Portfolio-Optimizer (greedy, DD-Constraint)
Mode 4: Equity-Kurven Chart (lokal + Telegram)

STRICT 70/30: All modes only use df[split_idx:] (the 30% OOS period).
The training data df[:split_idx] is NEVER used here.
"""
import json
import sys
from pathlib import Path
from typing import List, Dict

ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(ROOT / 'src'))

G  = '\033[0;32m'
Y  = '\033[1;33m'
C  = '\033[0;36m'
R  = '\033[0;31m'
B  = '\033[0;34m'
NC = '\033[0m'


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _load_configs() -> List[Dict]:
    """Load all config_*.json from strategy/configs/"""
    config_dir = ROOT / 'src' / 'probebot' / 'strategy' / 'configs'
    if not config_dir.exists():
        return []
    configs = []
    for p in sorted(config_dir.glob('config_*.json')):
        try:
            with open(p, encoding='utf-8') as f:
                c = json.load(f)
            c['_config_path'] = str(p)
            c['_config_name'] = p.stem
            configs.append(c)
        except Exception as e:
            print(f"  {R}Fehler beim Laden {p.name}: {e}{NC}")
    return configs


def _apply_capital_override(configs: List[Dict], total_override: float = None) -> List[Dict]:
    """
    Skaliert das Startkapital jeder Config proportional so, dass die Summe
    aller Configs total_override ergibt — die relativen Gewichte (aus dem
    bisherigen risk.start_capital je Config) bleiben erhalten. Da der
    Backtester die Positionsgroesse als fixen % von start_capital berechnet
    (siehe backtester.py), aendert eine proportionale Skalierung NUR die
    absoluten USDT-Betraege (Equity/PnL in USDT) — pnl_pct/win_rate/DD%
    bleiben identisch, da sich der Skalierungsfaktor in der %-Rechnung
    kuerzt. Ohne total_override werden die Configs unveraendert zurueckgegeben.
    """
    if total_override is None:
        return configs
    import copy
    weights = [c.get('risk', {}).get('start_capital', 100.0) for c in configs]
    weight_sum = sum(weights)
    out = []
    for c, w in zip(configs, weights):
        c2 = copy.deepcopy(c)
        c2.setdefault('risk', {})
        share = (w / weight_sum) if weight_sum > 0 else (1.0 / len(configs))
        c2['risk']['start_capital'] = round(total_override * share, 4)
        out.append(c2)
    return out


def _prompt_capital_override(configs: List[Dict]) -> List[Dict]:
    """Fragt ein Gesamt-Startkapital ab (Enter = Summe der Config-Werte,
    proportional aufgeteilt nach den bisherigen relativen Gewichten)."""
    default_total = sum(c.get('risk', {}).get('start_capital', 100.0) for c in configs)
    raw = input(f"  Gesamt-Startkapital USDT [Standard: {default_total:.0f} = Summe der Configs]: ").strip()
    if not raw:
        return configs
    try:
        total = float(raw)
    except ValueError:
        print(f"  {Y}Ungültige Eingabe, verwende Standard.{NC}")
        return configs
    return _apply_capital_override(configs, total)


def _rescale_result_capital(res: Dict, new_start_capital: float) -> Dict:
    """Skaliert ein bereits berechnetes Backtest-Ergebnis auf ein neues
    Startkapital, ohne den Backtest erneut laufen zu lassen — die
    Positionsgroesse ist ein fixer % von start_capital, daher skalieren alle
    USDT-Betraege linear mit (pnl_pct/win_rate/DD% bleiben unveraendert)."""
    import copy
    old_start = res['config'].get('risk', {}).get('start_capital', 100.0)
    if old_start <= 0:
        return res
    factor = new_start_capital / old_start
    res2 = copy.deepcopy(res)
    res2['config'].setdefault('risk', {})['start_capital'] = new_start_capital
    res2['end_capital'] = res['end_capital'] * factor
    for t in res2.get('trades', []):
        t['pnl'] = round(t['pnl'] * factor, 4)
        t['capital_after'] = round(t['capital_after'] * factor, 4)
    return res2


def _prompt_capital_override_results(results: List[Dict]) -> List[Dict]:
    """Wie _prompt_capital_override, aber fuer bereits fertige Backtest-
    Ergebnisse (Mode 3: Portfolio steht erst NACH der Greedy-Auswahl fest,
    ein erneuter Backtest waere unnoetig da nur USDT-Betraege skalieren)."""
    default_total = sum(r['config'].get('risk', {}).get('start_capital', 100.0) for r in results)
    raw = input(f"  Gesamt-Startkapital USDT [Standard: {default_total:.0f} = Summe der Configs]: ").strip()
    if not raw:
        return results
    try:
        total = float(raw)
    except ValueError:
        print(f"  {Y}Ungültige Eingabe, verwende Standard.{NC}")
        return results
    weights = [r['config'].get('risk', {}).get('start_capital', 100.0) for r in results]
    weight_sum = sum(weights)
    out = []
    for r, w in zip(results, weights):
        share = (w / weight_sum) if weight_sum > 0 else (1.0 / len(results))
        out.append(_rescale_result_capital(r, total * share))
    return out


def _load_oos_data(config: Dict, start_date: str = None, end_date: str = None):
    """
    Load the full feature df from parquet and slice to OOS period.

    Ohne start_date/end_date: Standard-OOS-Slice (df[split_idx:], die 30%
    die der Optimizer nie gesehen hat).

    Mit start_date/end_date: eigener Zeitraum (z.B. um zu pruefen ob ein
    Setup seit der Optimierung noch performt). Wird start_date VOR dem
    gespeicherten split_date gewaehlt, ist ein Teil des Zeitraums nicht
    mehr "out of sample" — intrusion_info gibt an wie viel.

    Returns df_oos, split_idx, intrusion_info (oder None, None, None bei Fehler).
    """
    import pandas as pd
    sym   = config['market']['symbol'].replace('/', '_').replace(':', '_')
    tf    = config['market']['timeframe']
    split_date = config.get('period', {}).get('split_date', '')

    data_path = ROOT / 'artifacts' / 'data' / f'data_{sym}_{tf}.parquet'
    if not data_path.exists():
        print(f"  {R}Daten nicht gefunden: {data_path.name}{NC}")
        print(f"  Erst run_pipeline.sh ausführen.")
        return None, None, None

    df = pd.read_parquet(data_path)

    # Find split index from split_date stored in config
    if not split_date:
        split_idx = int(len(df) * 0.70)
    else:
        mask = df['timestamp'].astype(str).str.startswith(split_date[:10])
        idxs = df[mask].index
        split_idx = int(idxs[0]) if len(idxs) > 0 else int(len(df) * 0.70)

    if not start_date:
        df_oos = df.iloc[split_idx:].copy().reset_index(drop=True)
        return df_oos, split_idx, None

    # ── Eigener Zeitraum ──────────────────────────────────────────────────
    tz = df['timestamp'].dt.tz
    start_ts = pd.Timestamp(start_date, tz=tz)
    mask = df['timestamp'] >= start_ts
    if end_date:
        end_ts = pd.Timestamp(end_date, tz=tz) + pd.Timedelta(days=1)
        mask &= df['timestamp'] < end_ts
    df_range = df[mask].copy().reset_index(drop=True)

    split_ts = df.iloc[split_idx]['timestamp'] if split_idx < len(df) else df['timestamp'].iloc[-1]
    intrusion_info = None
    if start_ts < split_ts:
        intruded = int((df_range['timestamp'] < split_ts).sum())
        total = len(df_range)
        intrusion_info = {
            'candles': intruded,
            'total': total,
            'pct': round(intruded / total * 100, 1) if total else 0.0,
            'split_date': str(split_ts)[:10],
        }

    return df_range, split_idx, intrusion_info


def _run_oos_backtest(config: Dict, start_date: str = None, end_date: str = None) -> Dict | None:
    """Run backtest on the OOS slice (Standard-30% oder eigener Zeitraum) fuer eine Config."""
    from probebot.analysis.backtester import run_backtest

    bot_spec_path = config.get('strategy', {}).get('bot_spec_path', '')
    if not Path(bot_spec_path).exists():
        # Try relative path from ROOT
        sym = config['market']['symbol'].replace('/', '_').replace(':', '_')
        tf  = config['market']['timeframe']
        bot_spec_path = str(ROOT / 'artifacts' / 'db' / f'bot_spec_{sym}_{tf}.json')

    if not Path(bot_spec_path).exists():
        print(f"  {R}bot_spec nicht gefunden: {bot_spec_path}{NC}")
        return None

    with open(bot_spec_path, encoding='utf-8') as f:
        bot_spec = json.load(f)

    entry_conditions = bot_spec.get('entry_conditions', {})
    tradeable        = config['strategy']['tradeable_types']
    params           = {**config.get('signal', {}), **config.get('risk', {})}
    start_capital    = config.get('risk', {}).get('start_capital', 100.0)

    df_oos, split_idx, intrusion_info = _load_oos_data(config, start_date, end_date)
    if df_oos is None or len(df_oos) == 0:
        return None

    result = run_backtest(df_oos, entry_conditions, tradeable, params, start_capital)
    result['config'] = config
    result['split_date'] = config.get('period', {}).get('split_date', '?')
    result['oos_end']    = config.get('period', {}).get('oos_end', '?')
    result['intrusion']  = intrusion_info
    if start_date:
        result['custom_start'] = start_date
        result['custom_end']   = end_date or str(df_oos['timestamp'].iloc[-1])[:10]
    return result


# ─── Mode 1: OOS-Backtest Tabelle ────────────────────────────────────────────

def mode_1_oos_table():
    from datetime import datetime

    configs = _load_configs()
    if not configs:
        print(f"\n  {R}Keine Configs gefunden.{NC}")
        print(f"  Erst run_pipeline.sh → Optimizer ausführen.")
        return

    print(f"\n{Y}Kapital:{NC}")
    configs = _prompt_capital_override(configs)

    print(f"\n{Y}Zeitraum:{NC}")
    print(f"  Enter = Standard (30%-OOS-Split je Config, wie im bot_spec gespeichert)")
    print(f"  Oder eigener Zeitraum (z.B. um zu prüfen ob ein Setup seit der")
    print(f"  Optimierung noch performt — gilt dann für alle Configs gleich)")
    start_input = input(f"  Start-Datum (YYYY-MM-DD) [Standard: automatisch je Config]: ").strip()
    custom_range = bool(start_input)
    end_input = None
    if custom_range:
        default_end = datetime.now().strftime('%Y-%m-%d')
        end_input = input(f"  End-Datum (YYYY-MM-DD) [Standard: {default_end} = heute]: ").strip()
        end_input = end_input or default_end

    print(f"\n{Y}{'─'*90}{NC}")
    if custom_range:
        print(f"{Y}  BACKTEST — eigener Zeitraum: {start_input} → {end_input}{NC}")
    else:
        print(f"{Y}  OOS-BACKTEST (30% Test-Periode — nie gesehen während Optimierung){NC}")
    print(f"{Y}{'─'*90}{NC}")
    print(f"  {'Config':<28}  {'Strategie':<10}  {'Trades':>6}  {'WinRate':>7}  "
          f"{'PnL%':>7}  {'MaxDD%':>7}  {'Sharpe':>7}  {'Periode'}")
    print(f"  {'─'*28}  {'─'*10}  {'─'*6}  {'─'*7}  {'─'*7}  {'─'*7}  {'─'*7}  {'─'*20}")

    results = []
    for cfg in configs:
        sym  = cfg['market']['symbol']
        tf   = cfg['market']['timeframe']
        name = f"{sym.split('/')[0]} {tf}"
        strat = cfg.get('strategy', {}).get('type', '?')
        print(f"  Berechne {name}...", end='\r', flush=True)
        res = _run_oos_backtest(cfg, start_input if custom_range else None, end_input if custom_range else None)
        if res is None:
            print(f"  {R}{name:<28}  FEHLER{NC}")
            continue
        results.append((name, strat, res))

    if not results:
        print(f"\n  {R}Keine Ergebnisse.{NC}")
        return

    results.sort(key=lambda x: x[2]['pnl_pct'], reverse=True)

    print()
    for name, strat, res in results:
        pnl_c = G if res['pnl_pct'] >= 0 else R
        dd_c  = R if res['max_drawdown'] > 20 else Y if res['max_drawdown'] > 10 else G
        wr_c  = G if res['win_rate'] >= 50 else Y if res['win_rate'] >= 40 else R
        if custom_range:
            period = f"{res['custom_start'][:7]} → {res['custom_end'][:7]}"
        else:
            period = f"{res['split_date'][:7]} → {res['oos_end'][:7]}"
        print(f"  {name:<28}  {strat:<10}  "
              f"{res['n_trades']:>6d}  "
              f"{wr_c}{res['win_rate']:>6.1f}%{NC}  "
              f"{pnl_c}{res['pnl_pct']:>+6.1f}%{NC}  "
              f"{dd_c}{res['max_drawdown']:>6.1f}%{NC}  "
              f"{res['sharpe']:>7.3f}  "
              f"{period}")
        if res.get('intrusion'):
            intr = res['intrusion']
            print(f"    {R}⚠ {intr['candles']}/{intr['total']} Kerzen ({intr['pct']}%) liegen VOR "
                  f"dem OOS-Split ({intr['split_date']}) — das sind Trainingsdaten, die der "
                  f"Optimizer bereits gesehen hat!{NC}")

    print(f"\n{Y}{'─'*90}{NC}")
    if custom_range:
        print(f"  Hinweis: eigener Zeitraum — ⚠-Zeilen markieren Configs bei denen ein Teil")
        print(f"  des gewählten Zeitraums NICHT mehr 'out of sample' ist (siehe oben).")
    else:
        print(f"  Hinweis: Diese Ergebnisse basieren auf Daten die der Optimizer NIE gesehen hat.")

    # Show detailed trades if user wants
    inp = input(f"\n  Trades einer Config anzeigen? (Name oder Enter zum Überspringen): ").strip()
    if inp:
        for name, strat, res in results:
            if inp.lower() in name.lower():
                _print_trade_detail(res['config'], res.get('trades', []))
                break


def _print_trade_detail(config: Dict, trades: List[Dict]):
    print(f"\n  {C}Trades Detail:{NC}")
    print(f"  {'Datum':<12}  {'Typ':<22}  {'Dir':>5}  {'PnL':>8}  {'Grund':<8}  {'Bars':>5}")
    print(f"  {'─'*12}  {'─'*22}  {'─'*5}  {'─'*8}  {'─'*8}  {'─'*5}")
    for t in trades:
        pnl_c = G if t['pnl'] > 0 else R
        print(f"  {str(t['entry_ts'])[:10]:<12}  "
              f"{t['move_type']:<22}  "
              f"{t['direction']:>5}  "
              f"{pnl_c}{t['pnl']:>+7.2f}{NC}  "
              f"{t['close_reason']:<8}  "
              f"{t['bars_held']:>5}")


# ─── Mode 2: Portfolio-Simulation ────────────────────────────────────────────

def mode_2_portfolio():
    configs = _load_configs()
    if not configs:
        print(f"\n  {R}Keine Configs.{NC}")
        return

    print(f"\n{Y}  Verfügbare Configs:{NC}")
    for i, cfg in enumerate(configs, 1):
        sym  = cfg['market']['symbol'].split('/')[0]
        tf   = cfg['market']['timeframe']
        strat = cfg.get('strategy', {}).get('type', '?')
        pnl   = cfg.get('_meta', {}).get('insample_pnl_pct', 0)
        print(f"  {i:2d}) {sym:<6} {tf:<4}  {strat:<10}  In-Sample: {pnl:+.1f}%")

    inp = input(f"\n  Welche kombinieren? (Nummern kommagetrennt, z.B. 1,3,5): ").strip()
    if not inp:
        return

    chosen_idxs = []
    for s in inp.split(','):
        try:
            idx = int(s.strip()) - 1
            if 0 <= idx < len(configs):
                chosen_idxs.append(idx)
        except ValueError:
            pass

    if not chosen_idxs:
        print(f"  {R}Ungültige Eingabe.{NC}")
        return

    chosen = [configs[i] for i in chosen_idxs]

    print(f"\n{Y}Kapital:{NC}")
    chosen = _prompt_capital_override(chosen)

    print(f"\n  Berechne OOS-Backtest für {len(chosen)} Configs...")

    all_results = []
    for cfg in chosen:
        res = _run_oos_backtest(cfg)
        if res:
            all_results.append(res)

    if not all_results:
        return

    # Combine equity curves (proportional to start_capital)
    _print_portfolio_summary(all_results)


def _print_portfolio_summary(results: List[Dict]):
    """Print combined portfolio stats."""
    total_start = sum(r['config'].get('risk', {}).get('start_capital', 100) for r in results)
    total_end   = sum(r['end_capital'] for r in results)
    combined_pnl = (total_end - total_start) / total_start * 100

    all_trades = []
    for r in results:
        for t in r.get('trades', []):
            all_trades.append(t)
    all_trades.sort(key=lambda t: t['entry_ts'])

    total_trades = sum(r['n_trades'] for r in results)
    wins = sum(1 for r in results for t in r.get('trades', []) if t['pnl'] > 0)
    wr   = wins / total_trades * 100 if total_trades > 0 else 0

    print(f"\n{Y}{'─'*60}{NC}")
    print(f"  PORTFOLIO (OOS 30%):")
    print(f"  Strategies:     {len(results)}")
    print(f"  Total Trades:   {total_trades}")
    print(f"  Win-Rate:       {wr:.1f}%")
    print(f"  Combined PnL:   {G if combined_pnl >= 0 else R}{combined_pnl:+.2f}%{NC}")
    print(f"  Einzel-PnL:")
    for r in results:
        sym = r['config']['market']['symbol'].split('/')[0]
        tf  = r['config']['market']['timeframe']
        pnl_c = G if r['pnl_pct'] >= 0 else R
        print(f"    {sym} {tf:<4}  {pnl_c}{r['pnl_pct']:+.1f}%{NC}  "
              f"({r['n_trades']} Trades, DD:{r['max_drawdown']:.1f}%)")
    print(f"{Y}{'─'*60}{NC}")


# ─── Mode 3: Auto-Portfolio-Optimizer ────────────────────────────────────────

def mode_3_auto_portfolio():
    """Greedy portfolio selection under drawdown constraint."""
    configs = _load_configs()
    if not configs:
        print(f"\n  {R}Keine Configs.{NC}")
        return

    try:
        max_dd = float(input(f"  Max. Portfolio-Drawdown % [Standard: 20]: ").strip() or "20")
    except ValueError:
        max_dd = 20.0

    print(f"\n  Berechne OOS-Backtest für alle {len(configs)} Configs...")
    all_results = []
    for cfg in configs:
        res = _run_oos_backtest(cfg)
        if res and res['n_trades'] >= 5:
            all_results.append(res)

    if not all_results:
        print(f"  {R}Keine verwertbaren Ergebnisse.{NC}")
        return

    # Sort by PnL descending
    all_results.sort(key=lambda r: r['pnl_pct'], reverse=True)

    # Greedy: add strategies while combined DD stays under limit
    portfolio = []
    for res in all_results:
        if res['max_drawdown'] > max_dd:
            continue  # individual DD already too high
        trial = portfolio + [res]
        combined_dd = _estimate_combined_dd(trial)
        if combined_dd <= max_dd:
            portfolio.append(res)

    if not portfolio:
        print(f"\n  {R}Kein Portfolio gefunden das DD ≤ {max_dd}% einhält.{NC}")
        print(f"  Tipp: max_dd erhöhen oder mehr Configs optimieren.")
        return

    print(f"\n{G}  Portfolio gefunden: {len(portfolio)} Strategien{NC}")

    print(f"\n{Y}Kapital:{NC}")
    portfolio = _prompt_capital_override_results(portfolio)

    _print_portfolio_summary(portfolio)

    # Optionally write to settings.json
    inp = input(f"\n  Portfolio in settings.json speichern? (j/n): ").strip().lower()
    if inp in ('j', 'y', 'ja', 'yes'):
        _write_portfolio_to_settings(portfolio)
        _generate_portfolio_equity_chart(portfolio)
        _generate_trades_excel(portfolio)


def _merge_portfolio_trades(portfolio: List[Dict]) -> List[Dict]:
    """
    Fuehrt alle Trades aller Configs chronologisch (nach entry_ts) zusammen und
    haengt an jeden Trade die laufende Portfolio-Gesamtequity an (Summe der
    zuletzt bekannten Einzel-Equity jeder Config zu diesem Zeitpunkt) — analog
    zu dnabots generate_portfolio_equity_chart()/generate_trades_excel().
    """
    all_trades = []
    for res in portfolio:
        cfg = res['config']
        sym = cfg['market']['symbol'].split('/')[0]
        tf  = cfg['market']['timeframe']
        start_cap = cfg.get('risk', {}).get('start_capital', 100.0)
        for t in res.get('trades', []):
            all_trades.append({**t, 'coin': sym, 'timeframe': tf, '_cfg_key': f"{sym}_{tf}",
                                '_start_capital': start_cap})
    all_trades.sort(key=lambda t: t['entry_ts'])

    last_equity = {t['_cfg_key']: t['_start_capital'] for t in all_trades}
    for t in all_trades:
        last_equity[t['_cfg_key']] = t['capital_after']
        t['portfolio_equity_after'] = round(sum(last_equity.values()), 4)

    return all_trades


def _generate_portfolio_equity_chart(portfolio: List[Dict]):
    """Kombinierter Portfolio-Equity-Chart (Plotly HTML), wie dnabot run_portfolio_optimizer.py."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"  {R}plotly nicht installiert — Chart übersprungen.{NC}")
        return

    all_trades = _merge_portfolio_trades(portfolio)
    if not all_trades:
        print(f"  {Y}Keine Trades — Chart übersprungen.{NC}")
        return

    total_start = sum(r['config'].get('risk', {}).get('start_capital', 100) for r in portfolio)

    PAIR_COLORS = ['#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6',
                   '#f97316', '#84cc16', '#06b6d4', '#a78bfa', '#eab308', '#22c55e']
    fig = make_subplots(specs=[[{'secondary_y': True}]])

    for idx, res in enumerate(portfolio):
        cfg = res['config']
        sym = cfg['market']['symbol'].split('/')[0]
        tf  = cfg['market']['timeframe']
        start_cap = cfg.get('risk', {}).get('start_capital', 100.0)
        pair_trades = sorted(res.get('trades', []), key=lambda t: t['entry_ts'])
        ptimes = [pair_trades[0]['entry_ts']] if pair_trades else []
        pvals  = [start_cap]
        for t in pair_trades:
            ptimes.append(t['close_ts'])
            pvals.append(t['capital_after'])
        fig.add_trace(go.Scatter(
            x=ptimes, y=pvals, mode='lines', name=f"{sym}/{tf}",
            line=dict(color=PAIR_COLORS[idx % len(PAIR_COLORS)], width=1), opacity=0.55,
        ), secondary_y=False)

    fig.add_hline(y=total_start, line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
                  annotation_text=f'Start {total_start:.0f} USDT', annotation_position='top left')

    eq_times = [all_trades[0]['entry_ts']] + [t['close_ts'] for t in all_trades]
    eq_vals  = [total_start] + [t['portfolio_equity_after'] for t in all_trades]

    entry_x, entry_y, entry_txt = [], [], []
    exit_tp_x, exit_tp_y   = [], []
    exit_sl_x, exit_sl_y   = [], []
    exit_liq_x, exit_liq_y = [], []
    exit_to_x, exit_to_y   = [], []
    for t in all_trades:
        eq_val = t['portfolio_equity_after']
        tip = f"{t['coin']} {t['timeframe']}<br>{t.get('move_type','')}<br>Equity: {eq_val:.2f} USDT"
        entry_x.append(t['close_ts']); entry_y.append(eq_val); entry_txt.append(tip)
        cr = t.get('close_reason', '')
        if cr == 'TP':
            exit_tp_x.append(t['close_ts']); exit_tp_y.append(eq_val)
        elif cr == 'SL':
            exit_sl_x.append(t['close_ts']); exit_sl_y.append(eq_val)
        elif cr == 'LIQ':
            exit_liq_x.append(t['close_ts']); exit_liq_y.append(eq_val)
        else:
            exit_to_x.append(t['close_ts']); exit_to_y.append(eq_val)

    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals, mode='lines', name='Portfolio Equity',
        line=dict(color='#2563eb', width=2), opacity=0.85,
    ), secondary_y=True)

    if exit_tp_x:
        fig.add_trace(go.Scatter(x=exit_tp_x, y=exit_tp_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=10, line=dict(width=1, color='#0e7490')),
            name='Exit TP ✓'), secondary_y=True)
    if exit_sl_x:
        fig.add_trace(go.Scatter(x=exit_sl_x, y=exit_sl_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=10, line=dict(width=2, color='#7f1d1d')),
            name='Exit SL ✗'), secondary_y=True)
    if exit_liq_x:
        fig.add_trace(go.Scatter(x=exit_liq_x, y=exit_liq_y, mode='markers',
            marker=dict(color='#ff00ff', symbol='diamond', size=12, line=dict(width=2, color='#7f007f')),
            name='Exit LIQ ⚠'), secondary_y=True)
    if exit_to_x:
        fig.add_trace(go.Scatter(x=exit_to_x, y=exit_to_y, mode='markers',
            marker=dict(color='#9ca3af', symbol='square', size=8),
            name='Exit Timeout'), secondary_y=True)

    n = len(all_trades)
    wins = sum(1 for t in all_trades if t['pnl'] > 0)
    wr = wins / n if n else 0.0
    final_equity = eq_vals[-1]
    pnl_pct = (final_equity - total_start) / total_start * 100 if total_start else 0.0
    sign = '+' if pnl_pct >= 0 else ''
    pairs_str = ', '.join(f"{r['config']['market']['symbol'].split('/')[0]}/{r['config']['market']['timeframe']}"
                           for r in portfolio)
    title = (f"probebot Portfolio — {len(portfolio)} Coins ({pairs_str}) | Trades: {n} | "
             f"WR: {wr:.1%} | PnL: {sign}{pnl_pct:.1f}% | Final Equity: {final_equity:.2f} USDT")

    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=750, hovermode='x unified', template='plotly_dark', dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
    )
    fig.update_yaxes(title_text='Einzel-Equity (USDT)', secondary_y=False, fixedrange=False)
    fig.update_yaxes(title_text='Portfolio-Equity (USDT)', secondary_y=True, fixedrange=False)

    charts_dir = ROOT / 'artifacts' / 'charts'
    charts_dir.mkdir(parents=True, exist_ok=True)
    out = charts_dir / 'probebot_portfolio_equity.html'
    fig.write_html(str(out))
    print(f"  {G}Portfolio-Chart erstellt: {out}{NC}")

    try:
        from probebot.utils.telegram import load_telegram_config, send_document
        secret_path = ROOT.parent / 'secret.json'
        if not secret_path.exists():
            secret_path = ROOT / 'secret.json'
        tg = load_telegram_config(str(secret_path))
        if tg.get('bot_token'):
            caption = (f"probebot Portfolio-Equity\n{len(portfolio)} Coins | "
                       f"PnL: {sign}{pnl_pct:.1f}% | Equity: {final_equity:.2f} USDT")
            if send_document(tg['bot_token'], tg['chat_id'], str(out), caption=caption):
                print(f"  {G}Via Telegram gesendet.{NC}")
            else:
                print(f"  {R}Telegram-Versand fehlgeschlagen.{NC}")
        else:
            print(f"  {Y}Telegram nicht konfiguriert.{NC}")
    except Exception as e:
        print(f"  {Y}Telegram-Versand übersprungen: {e}{NC}")


def _generate_trades_excel(portfolio: List[Dict]):
    """Excel-Tabelle aller Portfolio-Trades, wie dnabot run_portfolio_optimizer.py."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {Y}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return

    all_trades = _merge_portfolio_trades(portfolio)
    if not all_trades:
        print(f"  {Y}Keine Trades — Excel übersprungen.{NC}")
        return

    reason_map = {'TP': 'TP erreicht', 'SL': 'SL erreicht', 'LIQ': 'Liquidiert',
                  'TIMEOUT': 'Timeout', 'END': 'Periodenende'}
    rows = []
    for i, t in enumerate(all_trades, 1):
        rows.append({
            'Nr':                  i,
            'Datum':               t['entry_ts'][:16].replace('T', ' '),
            'Coin':                t['coin'],
            'Timeframe':           t['timeframe'],
            'Hebel':               f"{t.get('leverage', 0):.0f}x",
            'Richtung':            t['direction'],
            'Bewegungstyp':        t.get('move_type', ''),
            'Ergebnis':            reason_map.get(t.get('close_reason', ''), t.get('close_reason', '')),
            'PnL (USDT)':          t['pnl'],
            'Config-Kapital':      t['capital_after'],
            'Portfolio-Kapital':   t['portfolio_equity_after'],
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    win_fill     = PatternFill('solid', fgColor='D6F4DC')
    loss_fill    = PatternFill('solid', fgColor='FAD7D7')
    liq_fill     = PatternFill('solid', fgColor='F3D6FA')
    timeout_fill = PatternFill('solid', fgColor='FFF3CC')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                           top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    headers = list(rows[0].keys())
    col_widths = {'Nr': 6, 'Datum': 18, 'Coin': 10, 'Timeframe': 12, 'Hebel': 8, 'Richtung': 10,
                  'Bewegungstyp': 20, 'Ergebnis': 14, 'PnL (USDT)': 14,
                  'Config-Kapital': 16, 'Portfolio-Kapital': 18}

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        if row['Ergebnis'] == 'TP erreicht':
            fill = win_fill
        elif row['Ergebnis'] == 'SL erreicht':
            fill = loss_fill
        elif row['Ergebnis'] == 'Liquidiert':
            fill = liq_fill
        else:
            fill = timeout_fill if r_idx % 2 == 0 else alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('PnL (USDT)', 'Config-Kapital', 'Portfolio-Kapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    total_start = sum(r['config'].get('risk', {}).get('start_capital', 100) for r in portfolio)
    final_equity = rows[-1]['Portfolio-Kapital']
    pnl_pct = (final_equity - total_start) / total_start * 100 if total_start else 0.0
    wins = sum(1 for t in all_trades if t['pnl'] > 0)
    wr = wins / len(all_trades) * 100 if all_trades else 0.0

    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', len(all_trades)),
        ('Win-Rate', f"{wr:.1f}%"),
        ('PnL', f"{pnl_pct:+.1f}%"),
        ('Start-Kapital', f"{total_start:.2f} USDT"),
        ('Final Equity', f"{final_equity:.2f} USDT"),
        ('Coins', len(portfolio)),
    ]:
        ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=value)
        summary_row += 1

    charts_dir = ROOT / 'artifacts' / 'charts'
    charts_dir.mkdir(parents=True, exist_ok=True)
    out = charts_dir / 'probebot_trades.xlsx'
    wb.save(str(out))
    print(f"  {G}Excel-Tabelle erstellt: {out}{NC}")

    try:
        from probebot.utils.telegram import load_telegram_config, send_document
        secret_path = ROOT.parent / 'secret.json'
        if not secret_path.exists():
            secret_path = ROOT / 'secret.json'
        tg = load_telegram_config(str(secret_path))
        if tg.get('bot_token'):
            caption = (f"probebot Trades-Tabelle | {len(portfolio)} Coins | "
                       f"{len(all_trades)} Trades | WR: {wr:.1f}% | Final: {final_equity:.2f} USDT")
            if send_document(tg['bot_token'], tg['chat_id'], str(out), caption=caption):
                print(f"  {G}Via Telegram gesendet.{NC}")
            else:
                print(f"  {R}Telegram-Versand fehlgeschlagen.{NC}")
        else:
            print(f"  {Y}Telegram nicht konfiguriert.{NC}")
    except Exception as e:
        print(f"  {Y}Telegram-Versand übersprungen: {e}{NC}")


def _estimate_combined_dd(results: List[Dict]) -> float:
    """Simplified combined DD estimate: average of individual DDs."""
    if not results:
        return 0.0
    return sum(r['max_drawdown'] for r in results) / len(results) * 0.7


def _write_portfolio_to_settings(results: List[Dict]):
    settings_path = ROOT / 'settings.json'
    try:
        with open(settings_path, encoding='utf-8') as f:
            settings = json.load(f)
    except Exception:
        settings = {}

    active = []
    for res in results:
        cfg = res['config']
        active.append({
            'symbol':    cfg['market']['symbol'],
            'timeframe': cfg['market']['timeframe'],
            'strategy':  cfg.get('strategy', {}).get('type', 'HYBRID'),
            'active':    True,
        })

    # Write into live_trading_settings.active_strategies — that's where master_runner reads from
    if 'live_trading_settings' not in settings:
        settings['live_trading_settings'] = {}
    settings['live_trading_settings']['active_strategies'] = active
    settings_path.write_text(json.dumps(settings, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f"  {G}settings.json aktualisiert — {len(active)} aktive Strategien.{NC}")


# ─── Mode 4: Interaktiver Plotly Chart ───────────────────────────────────────

def mode_4_charts():
    """Interaktiver Plotly Chart (Candlestick + Trade-Marker + Equity) für OOS-Daten."""
    try:
        import plotly
    except ImportError:
        print(f"  {R}plotly nicht installiert. pip install plotly{NC}")
        return

    from probebot.analysis.interactive_chart import create_chart

    configs = _load_configs()
    if not configs:
        print(f"\n  {R}Keine Configs.{NC}")
        return

    print(f"\n{Y}  Verfügbare Configs:{NC}")
    for i, cfg in enumerate(configs, 1):
        sym   = cfg['market']['symbol'].split('/')[0]
        tf    = cfg['market']['timeframe']
        strat = cfg.get('strategy', {}).get('type', '?')
        pnl   = cfg.get('_meta', {}).get('insample_pnl_pct', 0)
        print(f"  {i:2d}) {sym:<6} {tf:<4}  {strat:<10}  In-Sample: {pnl:+.1f}%")

    raw = input(f"\n  Welche Config(s) als Chart? (Nummern kommagetrennt) [Standard: alle]: ").strip()

    if raw:
        chosen_idxs = []
        for s in raw.replace(',', ' ').split():
            try:
                idx = int(s.strip()) - 1
                if 0 <= idx < len(configs):
                    chosen_idxs.append(idx)
            except ValueError:
                pass
        chosen = [configs[i] for i in chosen_idxs]
    else:
        chosen = configs

    if not chosen:
        print(f"  {R}Ungültige Auswahl.{NC}")
        return

    print(f"\n{Y}Kapital:{NC}")
    chosen = _prompt_capital_override(chosen)

    from datetime import datetime
    print(f"\n{Y}Zeitraum:{NC}")
    print(f"  Enter = Standard (30%-OOS-Split je Config, wie im bot_spec gespeichert)")
    print(f"  Oder eigener Zeitraum (z.B. um zu prüfen ob ein Setup seit der")
    print(f"  Optimierung noch performt — gilt dann für alle Configs gleich)")
    start_input = input(f"  Start-Datum (YYYY-MM-DD) [Standard: automatisch je Config]: ").strip()
    custom_range = bool(start_input)
    end_input = None
    if custom_range:
        default_end = datetime.now().strftime('%Y-%m-%d')
        end_input = input(f"  End-Datum (YYYY-MM-DD) [Standard: {default_end} = heute]: ").strip()
        end_input = end_input or default_end

    charts_dir = ROOT / 'artifacts' / 'charts'
    charts_dir.mkdir(parents=True, exist_ok=True)
    generated  = []

    for cfg in chosen:
        sym  = cfg['market']['symbol']
        tf   = cfg['market']['timeframe']
        name = f"{sym.split('/')[0]} {tf}"
        print(f"\n  Berechne OOS-Backtest: {name}...")

        s = start_input if custom_range else None
        e = end_input if custom_range else None
        res = _run_oos_backtest(cfg, s, e)
        if res is None or not res.get('trades'):
            print(f"  {R}{name}: Keine OOS-Trades.{NC}")
            continue

        df_oos, _, intrusion_info = _load_oos_data(cfg, s, e)
        if df_oos is None:
            continue

        start_cap = cfg.get('risk', {}).get('start_capital', 100.0)
        print(f"  {G}{res['n_trades']} Trades | WR:{res['win_rate']:.1f}% | "
              f"PnL:{res['pnl_pct']:+.1f}% | DD:{res['max_drawdown']:.1f}%{NC}")
        if intrusion_info:
            print(f"  {R}⚠ {intrusion_info['candles']}/{intrusion_info['total']} Kerzen "
                  f"({intrusion_info['pct']}%) liegen VOR dem OOS-Split "
                  f"({intrusion_info['split_date']}) — das sind Trainingsdaten, die der "
                  f"Optimizer bereits gesehen hat! Wird im Chart rot markiert.{NC}")
        print(f"  Erstelle Chart...")

        fig = create_chart(sym, tf, df_oos, res['trades'], res, start_cap, intrusion=intrusion_info)
        if fig is None:
            continue

        safe = f"{sym.replace('/', '').replace(':', '')}_{tf}"
        out  = charts_dir / f'probebot_{safe}.html'
        fig.write_html(str(out))
        print(f"  {G}Chart gespeichert: {out}{NC}")
        generated.append(str(out))

    if generated:
        print(f"\n  {G}{len(generated)} Chart(s) generiert.{NC}")

    # Telegram
    try:
        from probebot.utils.telegram import load_telegram_config, send_document
        secret_path = ROOT.parent / 'secret.json'
        if not secret_path.exists():
            secret_path = ROOT / 'secret.json'
        tg = load_telegram_config(str(secret_path))
        if tg.get('bot_token') and generated:
            inp = input("  Per Telegram senden? (j/n): ").strip().lower()
            if inp in ('j', 'y', 'ja'):
                for path in generated:
                    send_document(tg['bot_token'], tg['chat_id'], path,
                                  caption=f"Probebot Chart: {Path(path).stem}")
                print(f"  {G}Telegram gesendet.{NC}")
    except Exception:
        pass


# ─── Main entry point ────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', type=int, default=None)
    args = parser.parse_args()

    if args.mode is None:
        print(f"\n{B}{'='*60}{NC}")
        print(f"       probebot — Trading Bot Evaluation")
        print(f"{B}{'='*60}{NC}")
        print(f"\n{Y}Modus wählen:{NC}")
        print(f"  {G}1{NC}) OOS-Backtest (30% Test-Periode) aller Configs")
        print(f"  {G}2{NC}) Portfolio-Simulation (mehrere Configs kombinieren)")
        print(f"  {G}3{NC}) Auto-Portfolio-Optimizer (greedy, DD-Constraint)")
        print(f"  {G}4{NC}) Equity-Chart erstellen + Telegram")
        mode = input(f"\nAuswahl (1-4) [Standard: 1]: ").strip() or "1"
    else:
        mode = str(args.mode)

    mode = mode.strip()
    if mode == '1':
        mode_1_oos_table()
    elif mode == '2':
        mode_2_portfolio()
    elif mode == '3':
        mode_3_auto_portfolio()
    elif mode == '4':
        mode_4_charts()
    else:
        print(f"  {R}Ungültiger Modus: {mode}{NC}")


if __name__ == '__main__':
    main()
